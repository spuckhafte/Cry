import random
from datetime import datetime
import json
from bot_bio import Behaviour
import openpyxl as xl
from cryptography.fernet import Fernet
import hashlib

SIGN = Behaviour.hash_sign
DEF_AMOUNT = Behaviour.default_amount

key = Fernet.generate_key()
fernet = Fernet(key)

try:
    with open('functions/ledger.xlsx', 'rb') as outfile:
        ledger_wb = xl.load_workbook(outfile)
except FileNotFoundError:
    with open('ledger.xlsx', 'rb') as outfile:
        ledger_wb = xl.load_workbook(outfile)
ledger_sheet = ledger_wb['Sheet1']

# 1 CryCoin = $100 = Rs.7601 # LOL

# value of the entity
info = {
    'Cr.1': '$.100',
    'Format': '1/10,?:CryCoin(1812210)/0000000000000000000000000000000000000000000000000000000000000000/1157551/--2005ca051baeb3b0557ee056c28ec3f0e76e25bf3a2fafd14227b6a3c2bae2ae'
}


class EncDeEnc:
    def __init__(self, deEncrypted='', Encrypted=b''):
        self.enc = Encrypted
        self.deEnc = deEncrypted

    def value_encrypt(self):
        enc_value = fernet.encrypt(self.deEnc.encode())
        return enc_value

    def value_decrypt(self):
        de_enc_value = fernet.decrypt(self.enc).decode()
        return de_enc_value

    def hash_encrypt(self):
        enc = hashlib.sha256(self.deEnc.encode()).hexdigest()
        return enc


# print(EncDeEnc(deEncrypted='hesoyam').hash_encrypt())


def mine(transaction_data):
    array_transaction_data = transaction_data.split('/')
    if len(array_transaction_data) > 3:
        array_transaction_data.pop()
    i = 0

    while True:
        if len(array_transaction_data) > 3:
            array_transaction_data.pop()

        array_transaction_data.append(str(i))
        new_transaction_data = '/'.join(array_transaction_data)
        encryption_on_data = EncDeEnc(deEncrypted=new_transaction_data).hash_encrypt()

        if encryption_on_data.startswith(SIGN):
            new_transaction_data += f'/~~{encryption_on_data}'
            return [new_transaction_data, encryption_on_data]
        else:
            i += 1

# print(mine('14/1e-15,?:argøn#0699(7120227)/2005cc6fa9afa76cf841925687839c96f2ba8ef28cc89dc2f3aad57cecfa53cd/nonce'))

def max_row():
    return ledger_sheet.max_row - 1


def previous_hash():
    row = max_row()
    prev = ledger_sheet.cell(row+1, 8).value if row != 0 else Behaviour.default_hash
    return prev


def genesis_block():
    day, month, year, hour = str(datetime.now().day), str(datetime.now().month), str(datetime.now().year), str(
        datetime.now().hour)
    datetime_str = day + month + year + hour

    from_ = '?'

    fac = Behaviour.cry_factor
    with open('members.json', 'r') as infile:
        infile = json.load(infile)
    members = [user['username'] for user in infile['users']]
    cries = [float(user['cries']) * fac if float(user['cries']) != 0 else 1/fac for user in infile['users']]

    amount = float(DEF_AMOUNT) if max_row() == 0 else float(1/fac)
    to = random.choices(members, cries)[0]
    prev_hash = previous_hash()

    details = [str(max_row() + 1), amount, from_, to, datetime_str, f'{prev_hash}', 'nonce']
    raw_transaction_string = f"{details[0]}/{details[1]},{details[2]}:{details[3]}({details[4]})/{details[5]}/{details[6]}"

    with open('members.json', 'r') as file:
        hash_trans_string = json.load(file)

    hash_trans_string['current-unmined-string'].append(raw_transaction_string)
    var = datetime.now()
    hash_trans_string['last-activity'] = f'{var.year}/{var.month}/{var.day} {var.hour}:{var.minute}:{var.second}'

    with open('members.json', 'w') as file:
        json.dump(hash_trans_string, file, indent=4)

    export_data = {
        'cries': amount,
        'to': to,
        'from': from_
    }

    return export_data


def sync_xl(data):
    dumping_data = [max_row() + 1, data['amount'], data['from'], data['to'], data['event'], data['last-hash'],
                    f"{data['string']}/~~{data['hash']}", data['hash']]

    row = max_row()+2
    for index in range(1, len(dumping_data)+1):
        ledger_sheet.cell(row, index).value = dumping_data[index-1]
    ledger_wb.save('./functions/ledger.xlsx')


async def check_hash_availiblity(hash_check):
    available = True
    for i in range(1, max_row()+1):
        if ledger_sheet.cell(i, 8).value == hash_check:
            available = False
            break
    return available


# print(mine('4/3.749659198247581e-08,spuckhafte_ferwirklung#7109:ARGØN#0699(2312202117)/2005cc8414bd4d358f7d1fdc77fd04d17ac3f6ffae77a2a31a3c49b7b71890bb/nonce'))
